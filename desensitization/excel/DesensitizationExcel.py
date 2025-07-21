import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
# excel批量脱敏

def desensitize_name(name):
    """脱敏姓名：不超过三个字的姓名，屏蔽第一个字；超过三个字，屏蔽第一个、第二个字，用*代替"""
    if pd.isna(name) or not str(name).strip():
        return name
    name_str = str(name)
    if len(name_str) <= 3:
        return '*' + name_str[1::]
    return '*' * 2 + name_str[2::]


def desensitize_id(id_number):
    """脱敏身份证号：屏蔽后6位"""
    if pd.isna(id_number) or not str(id_number).strip():
        return id_number
    id_str = str(id_number)
    if len(id_str) <= 6:
        return '*' * len(id_str)
    return id_str[:12] + '*' * 6


def desensitize_bank_card(card_number):
    """脱敏银行卡号：保留前6位和后4位"""
    if pd.isna(card_number) or not str(card_number).strip():
        return card_number
    card_str = str(card_number)
    if len(card_str) <= 10:
        return '*' * len(card_str)
    return card_str[:6] + '*' * (len(card_str) - 10) + card_str[-4:]


def process_excel_file(input_path, output_path, columns_config):
    """
    处理单个Excel文件
    :param input_path: 输入文件路径
    :param output_path: 输出文件路径
    :param columns_config: 列配置，格式为 {'列名': '脱敏类型'}
    """
    # 读取Excel文件
    df = pd.read_excel(input_path)

    # 应用脱敏规则
    for column, desensitize_type in columns_config.items():
        if column in df.columns:
            if desensitize_type == 'name':
                df[column] = df[column].apply(desensitize_name)
            elif desensitize_type == 'id':
                df[column] = df[column].apply(desensitize_id)
            elif desensitize_type == 'bank_card':
                df[column] = df[column].apply(desensitize_bank_card)

    # 保存处理后的文件
    if input_path.endswith('.xlsx'):
        # 对于.xlsx文件，使用openpyxl保留原格式
        wb = load_workbook(input_path)
        ws = wb.active

        # 清空原有数据
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None

        # 写入新数据
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        wb.save(output_path)
    else:
        # 对于其他格式，直接使用pandas保存
        df.to_excel(output_path, index=False)


def batch_process_excel(input_folder, output_folder, columns_config):
    """
    批量处理Excel文件
    :param input_folder: 输入文件夹路径
    :param output_folder: 输出文件夹路径
    :param columns_config: 列配置，格式为 {'列名': '脱敏类型'}
    """
    # 确保输出文件夹存在
    os.makedirs(output_folder, exist_ok=True)

    # 遍历输入文件夹中的所有Excel文件
    for filename in os.listdir(input_folder):
        if filename.endswith(('.xls', '.xlsx')):
            input_path = os.path.join(input_folder, filename)
            output_path = os.path.join(output_folder, filename)

            print(f"正在处理文件: {filename}")
            try:
                process_excel_file(input_path, output_path, columns_config)
                print(f"文件处理完成，已保存到: {output_path}")
            except Exception as e:
                print(f"处理文件 {filename} 时出错: {str(e)}")


if __name__ == "__main__":
    # 配置输入输出文件夹
    input_folder = "D:\\tmp\\input"  # 存放待处理Excel文件的文件夹
    output_folder = "D:\\tmp\\output"  # 处理后的文件输出文件夹

    # 配置需要脱敏的列及脱敏类型
    # 格式: {'列名': '脱敏类型'}，脱敏类型可以是 'name', 'id', 'bank_card'
    columns_config = {
        '姓名': 'name',
        '身份证号': 'id',
        '银行卡号': 'bank_card',
        # 可以添加更多列配置
    }

    # 执行批量处理
    batch_process_excel(input_folder, output_folder, columns_config)