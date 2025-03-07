import os
import sys

import mysql.connector
import paramiko
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from scp import SCPClient

# 该脚本，通过读取excel中档案号（D列），通过MySQL查询该档案号下资料，
# 并按照E列，作为文件目录命名，通过SCP方式下载资料

# excel绝对路径
xlsx_path = r'D:\cs_download\审核底稿-下载.xlsx'
# sheet名
sheet_name = 'Sheet1'
# 读取excel中指定sheet数据
wb = load_workbook(xlsx_path, read_only=True)
ws = wb[sheet_name]
row = ws.max_row
data = {}
for i in range(1, row):
    # 档案号
    key = ws.cell(i + 1, column_index_from_string('D')).value
    # 文件夹命名
    name = ws.cell(i + 1, column_index_from_string('E')).value
    data[i + 1] = (key, name)
wb.close()

dataSize = len(data)
if dataSize > 200:
    print(f"单次最多下载200个用户资料，当前：{dataSize}，程序退出。")
    sys.exit()

# 连接配置信息
config = {
    'user': 'root',  # 你的MySQL用户名
    'password': 'taxBook@2021',  # 你的MySQL密码
    'host': '192.168.1.106',  # 数据库服务器地址
    'port': 3308,  # 端口
    'database': 'zzdb',  # 数据库名
    'raise_on_warnings': True
}

# 建立连接
try:
    for itemId, value in data.values():
        conn = mysql.connector.connect(**config)
        # 创建游标对象
        cursor = conn.cursor()
        # 执行SQL查询 并 拼接参数
        query = "SELECT filename FROM lose_mod_file f inner join lose l on l.id = f.loseid " \
                "where l.zongxuhao = %s "
        cursor.execute(query, (itemId,))
        # 获取查询结果
        results = cursor.fetchall()
        # 打印查询结果
        print(f'查询当前档案号：{itemId}，共{len(results)}个文件，开始下载')

        # 服务器连接配置
        linux_hostname = '192.168.1.109'
        linux_port = 22
        linux_username = 'root'
        linux_password = 'dazu123.com'

        # 创建SSH客户端
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        # 连接到服务器
        ssh.connect(linux_hostname, linux_port, linux_username, linux_password)

        # 目标文件夹路径
        source_path = '/dazu/www/zssite/uptownfile'
        target_path = '/target-dir/' + value

        # 确保目标目录存在
        if not os.path.exists(target_path):
            os.makedirs(target_path)

        # 创建SCP客户端
        scp = SCPClient(ssh.get_transport())
        # 复制文件
        for file_tuple in results:
            file_name = file_tuple[0]
            source_file = f'{source_path}/{file_name}'
            target_file = f'{target_path}/{file_name}'
            try:
                scp.get(source_file, target_file)
                print(f'已拷贝文件：{file_name}')
            except Exception as e:
                print(f'****** 未找到文件 ******：{file_name}')
        print(f'{itemId}对应文件下载完成,目录：{target_path}')
        print('*' * 40)
        # 关闭连接
        scp.close()
        ssh.close()
    print(f"文件均下载完成，档案号：{data}")
except mysql.connector.Error as err:
    print(f"Error: {err}")

finally:
    # 关闭游标和连接
    if conn.is_connected():
        cursor.close()
        conn.close()
        print("MySQL connection is closed")
