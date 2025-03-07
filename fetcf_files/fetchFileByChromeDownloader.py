import os
import time
from tqdm import tqdm
import pyautogui

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException, UnexpectedAlertPresentException

path = r'D:\cs_download\dir'
xlsx_path = r'D:\cs_download\审核底稿-下载.xlsx'
sheet_name = 'Sheet1'
txt_path = rf'D:\cs_download\name.txt'
fin = int(input('fin = '))  # 174

os.chdir(path)
files = sorted(os.listdir(os.getcwd()), key=os.path.getmtime)
file_num = len(files)

wb = load_workbook(xlsx_path, read_only=True)
ws = wb[sheet_name]
row = ws.max_row
# row = 231
data = {}
for i in range(1, row):
    key = ws.cell(i + 1, column_index_from_string('D')).value
    name = ws.cell(i + 1, column_index_from_string('E')).value 
    data[i + 1] = (key, name)
wb.close()

options = Options()
pref = {"download.default_directory": path}
options.add_experimental_option("prefs", pref)


###options.add_argument("--headless")
###driver_width, driver_height = pyautogui.size()
###options.add_argument('--window-size=%sx%s' % (driver_width, driver_height))

driver = webdriver.Chrome(options=options)
driver.implicitly_wait(5)

driver.maximize_window()

driver.get("http://192.168.1.109:8090/")
driver.find_element(By.NAME, "username").send_keys("22125861")
driver.find_element(By.NAME, "password").send_keys("zz123456")
driver.find_element(By.CLASS_NAME, "login-btn").click()


def wait():
    os.chdir(path)
    while True:
        new_files = sorted(os.listdir(os.getcwd()), key=os.path.getmtime)
        if len(new_files) == file_num:
            time.sleep(1)
        else:
            newest = new_files[-1]

            if ".crdownload" in newest or ".tmp" in newest:
                time.sleep(1)
            else:
                return newest


tag = True
index = fin
while True:
    for num in tqdm(range(fin, row)):
        index += 1
        key, name = data[num + 1]
        try:
            url = "http://192.168.1.109:8090/?keyword=+" + key + "+"
        ###输入档案号
            driver.get(url)
            time.sleep(1)

            driver.find_element(By.XPATH, "//tr[@class='tr_1ine openDetail']").click()
          
            time.sleep(1)
            driver.switch_to.window(driver.window_handles[1])

            driver.find_element(By.XPATH, "//*[@id='report']//div[@class='foot_button']/div/input").click()
            time.sleep(1)
            driver.find_element(By.XPATH, "//*[@id='report']//div[@class='foot_button']/a[1]").click()
            time.sleep(3) #

            try:
                driver.close()
            except UnexpectedAlertPresentException:
                tag = False
                txt = open(txt_path, 'a')
                txt.writelines(f'{num + 1} {key} {name} FFFF00\n')
                txt.close()
                print(f'\n{num + 1}: NoFileError! ')
            else:
                tag = True
                driver.switch_to.window(driver.window_handles[0])
                time.sleep(1)

                files = sorted(os.listdir(os.getcwd()), key=os.path.getmtime)
                if len(files) == file_num:
                    driver.find_element(By.XPATH, "//tr[@class='tr_1ine openDetail']").click()
                    time.sleep(1)
                    driver.switch_to.window(driver.window_handles[1])

                    driver.find_element(By.XPATH, "//*[@id='report']//div[@class='foot_button']/div/input").click()
                    time.sleep(1)
                    driver.find_element(By.XPATH, "//*[@id='report']//div[@class='foot_button']/a[1]").click()
                    time.sleep(3) #

                old_name = wait()
                if old_name is not None:
                    files = sorted(os.listdir(os.getcwd()), key=os.path.getmtime)
                    file_num = len(files)
                    txt = open(txt_path, 'a')
                    txt.writelines(f'{old_name} {name}.zip\n')
                    txt.close()

        except Exception as e:
            if type(e) == NoSuchElementException:
                color = 'FF0000'
            else:
                color = '00B0F0'
            txt = open(txt_path, 'a')
            txt.writelines(f'{num + 1} {key} {name} {color}\n')
            txt.close()

            tag = False
            error = str(type(e)).strip("<class '").strip("'>").strip("__main__.").split('.')[-1]
            print(f'\n{num + 1}: {error}! ')

        if len(driver.window_handles) > 1:
            driver.close()
        driver.switch_to.window(driver.window_handles[0])
        time.sleep(1)

        if not tag:
            break

    fin = index
    if index == row:
        break

wb.close()
driver.quit()
